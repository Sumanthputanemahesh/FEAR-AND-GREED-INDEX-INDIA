"""
India Fear & Greed Index — daily Excel export.

Run after build_india_fgi.py / publish_latest.py:  python3 scripts/export_excel.py

Writes docs/data/India_FearGreed_History.xlsx, regenerated in full every day
from india_fgi_core.csv and india_fgi_extended.csv — the same official, close-based
daily record that feeds the website's history chart. This file is NEVER
touched by scripts/intraday_update.py; it only ever reflects settled,
end-of-day scores.

Sheets:
  Core     - 5-component composite, full history (2020-06-11 onward)
  Extended - 6-component composite (adds Put/Call), shorter history
             (bounded by NSE bhavcopy's format only existing from mid-2024)
  Read Me  - what this file is, what each sheet means, last-updated stamp
"""

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "docs" / "data"
CORE_CSV = DATA_DIR / "india_fgi_core.csv"
EXT_CSV = DATA_DIR / "india_fgi_extended.csv"
OUT_XLSX = DATA_DIR / "India_FearGreed_History.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ZONE_FILLS = {
    "Extreme Fear": PatternFill("solid", fgColor="C0504D"),
    "Fear": PatternFill("solid", fgColor="E6B8B7"),
    "Neutral": PatternFill("solid", fgColor="FFF2CC"),
    "Greed": PatternFill("solid", fgColor="C6E0B4"),
    "Extreme Greed": PatternFill("solid", fgColor="548235"),
}


def load(path):
    if not path.exists():
        return None
    df = pd.read_csv(path, index_col=0, parse_dates=True)
    return df if not df.empty else None


def style_sheet(ws, df, zone_col_name="zone"):
    ws.freeze_panes = "A2"
    headers = ["Date"] + list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    zone_col_idx = headers.index(zone_col_name) + 1 if zone_col_name in headers else None
    for row_idx in range(2, ws.max_row + 1):
        if zone_col_idx:
            zone_val = ws.cell(row=row_idx, column=zone_col_idx).value
            fill = ZONE_FILLS.get(zone_val)
            if fill:
                ws.cell(row=row_idx, column=zone_col_idx).fill = fill

    for col_idx, header in enumerate(headers, start=1):
        width = max(12, len(header) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    core = load(CORE_CSV)
    ext = load(EXT_CSV)
    if core is None and ext is None:
        print("FATAL: no core or extended data to export")
        return

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        if core is not None:
            out = core.reset_index().rename(columns={core.index.name or "index": "Date"})
            out["Date"] = out["Date"].dt.strftime("%Y-%m-%d")
            out.to_excel(writer, sheet_name="Core", index=False)

        if ext is not None:
            out = ext.reset_index().rename(columns={ext.index.name or "index": "Date"})
            out["Date"] = out["Date"].dt.strftime("%Y-%m-%d")
            out.to_excel(writer, sheet_name="Extended", index=False)

        readme_rows = [
            ["India Fear & Greed Index — Full History"],
            [""],
            ["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
            ["Core sheet", f"{len(core)} days, {core.index.min().date()} to {core.index.max().date()}" if core is not None else "unavailable"],
            ["Extended sheet", f"{len(ext)} days, {ext.index.min().date()} to {ext.index.max().date()}" if ext is not None else "unavailable"],
            [""],
            ["Core", "5 components (Momentum, Volatility, Strength, Breadth, Safe Haven). "
                     "Longest, most reliable history — use this for any analysis."],
            ["Extended", "Adds Credit Spread and/or Put/Call where their shorter, less reliable "
                         "feeds allow it. A given date's Core and Extended scores are NOT directly "
                         "comparable — they're computed from different component sets."],
            [""],
            ["Methodology", "Each component is a rolling 252-trading-day percentile rank "
                             "((count of values below current) / 251 * 100). Volatility, Put/Call, "
                             "and Credit Spread are inverted. Composite = equal-weighted mean."],
            ["Zones", "0-25 Extreme Fear · 26-45 Fear · 46-55 Neutral · 56-75 Greed · 76-100 Extreme Greed"],
            [""],
            ["Source", "https://github.com/Sumanthputanemahesh/FEAR-AND-GREED-INDEX-INDIA"],
            ["Live site", "https://sumanthputanemahesh.github.io/FEAR-AND-GREED-INDEX-INDIA/"],
            [""],
            ["Note", "This file reflects official, close-based daily scores only. It is "
                     "regenerated in full every day and is never touched by the site's hourly "
                     "intraday live estimate."],
        ]
        readme_df = pd.DataFrame(readme_rows)
        readme_df.to_excel(writer, sheet_name="Read Me", index=False, header=False)

        for sheet_name, df in [("Core", core), ("Extended", ext)]:
            if df is None:
                continue
            ws = writer.sheets[sheet_name]
            style_sheet(ws, df)

        ws = writer.sheets["Read Me"]
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 100
        ws["A1"].font = Font(bold=True, size=14)

    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
